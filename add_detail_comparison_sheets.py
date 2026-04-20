from __future__ import annotations

import argparse
import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


COMPARE_FIELDS = [
    "office",
    "ramo",
    "poliza",
    "document",
    "issue_date",
    "description",
    "prima_neta",
    "pct_comision",
    "comision",
    "descuento",
    "raw_line",
]


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
DIFF_FILL = PatternFill("solid", fgColor="FFF2CC")
META_FILL = PatternFill("solid", fgColor="E2F0D9")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Agrega hojas para comparar detalle_comisiones digital vs scan lado a lado."
    )
    parser.add_argument(
        "--workbook",
        default="output/positiva_comisiones_all.xlsx",
        help="Workbook base que contiene la hoja detalle_comisiones.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Ruta de salida. Si no se indica, intenta guardar sobre el mismo workbook.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    workbook_path = Path(args.workbook)
    output_path = Path(args.output) if args.output else workbook_path
    append_comparison_sheets(workbook_path, output_path)
    print(output_path.resolve())
    return 0


def append_comparison_sheets(workbook_path: Path, output_path: Path | None = None) -> None:
    output_path = output_path or workbook_path
    detail_df = pd.read_excel(workbook_path, sheet_name="detalle_comisiones")
    detail_df["row_index"] = detail_df.groupby("source_file").cumcount() + 1

    workbook = load_workbook(workbook_path)
    digital_files = [
        source_file
        for source_file in detail_df["source_file"].drop_duplicates().tolist()
        if not str(source_file).endswith("_scan.pdf")
    ]

    for digital_file in digital_files:
        scan_file = _build_scan_name(digital_file)
        if scan_file not in set(detail_df["source_file"]):
            continue
        sheet_name = _build_sheet_name(digital_file, detail_df)
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        worksheet = workbook.create_sheet(title=sheet_name)
        digital_rows = detail_df.loc[detail_df["source_file"] == digital_file].copy()
        scan_rows = detail_df.loc[detail_df["source_file"] == scan_file].copy()
        _write_comparison_sheet(worksheet, digital_rows, scan_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _build_scan_name(digital_file: str) -> str:
    path = Path(digital_file)
    return f"{path.stem}_scan{path.suffix}"


def _build_sheet_name(digital_file: str, detail_df: pd.DataFrame) -> str:
    boleta = str(
        detail_df.loc[detail_df["source_file"] == digital_file, "boleta_number"].iloc[0]
    )
    entity = str(detail_df.loc[detail_df["source_file"] == digital_file, "entity"].iloc[0])
    entity_token = entity.replace("POSITIVA ", "").replace(" ", "_")[:10]
    return f"CMP_{entity_token}_{boleta}"[:31]


def _write_comparison_sheet(worksheet, digital_rows: pd.DataFrame, scan_rows: pd.DataFrame) -> None:
    digital_rows = digital_rows.sort_values("row_index").reset_index(drop=True)
    scan_rows = scan_rows.sort_values("row_index").reset_index(drop=True)

    digital_file = str(digital_rows["source_file"].iloc[0])
    scan_file = str(scan_rows["source_file"].iloc[0])
    entity = str(digital_rows["entity"].iloc[0])
    boleta = str(digital_rows["boleta_number"].iloc[0])

    worksheet["A1"] = "Digital"
    worksheet["B1"] = digital_file
    worksheet["D1"] = "Scan"
    worksheet["E1"] = scan_file
    worksheet["A2"] = "Entidad"
    worksheet["B2"] = entity
    worksheet["D2"] = "Boleta"
    worksheet["E2"] = boleta

    for cell_ref in ("A1", "D1", "A2", "D2"):
        worksheet[cell_ref].font = Font(bold=True)
        worksheet[cell_ref].fill = META_FILL

    header_row = 4
    worksheet.cell(header_row, 1, "fila")
    current_column = 2
    for field in COMPARE_FIELDS:
        worksheet.cell(header_row, current_column, f"{field}_digital")
        worksheet.cell(header_row, current_column + 1, f"{field}_scan")
        current_column += 2
    worksheet.cell(header_row, current_column, "campos_distintos")

    for cell in worksheet[header_row]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    max_rows = max(len(digital_rows), len(scan_rows))
    for row_offset in range(max_rows):
        excel_row = header_row + 1 + row_offset
        worksheet.cell(excel_row, 1, row_offset + 1)
        digital_record = _row_to_dict(digital_rows, row_offset)
        scan_record = _row_to_dict(scan_rows, row_offset)
        current_column = 2
        diff_fields: list[str] = []

        for field in COMPARE_FIELDS:
            digital_value = digital_record.get(field)
            scan_value = scan_record.get(field)
            worksheet.cell(excel_row, current_column, digital_value)
            worksheet.cell(excel_row, current_column + 1, scan_value)
            if _normalize_for_compare(digital_value) != _normalize_for_compare(scan_value):
                diff_fields.append(field)
                worksheet.cell(excel_row, current_column).fill = DIFF_FILL
                worksheet.cell(excel_row, current_column + 1).fill = DIFF_FILL
            current_column += 2

        worksheet.cell(excel_row, current_column, " | ".join(diff_fields))

    worksheet.freeze_panes = "A5"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_columns(worksheet)


def _row_to_dict(frame: pd.DataFrame, row_offset: int) -> dict:
    if row_offset >= len(frame):
        return {}
    return frame.iloc[row_offset].to_dict()


def _normalize_for_compare(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, float):
        return f"{value:.2f}"
    text = str(value).strip()
    return " ".join(text.split())


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        max_length = max((len(value) for value in values), default=0)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 45)


if __name__ == "__main__":
    raise SystemExit(main())
